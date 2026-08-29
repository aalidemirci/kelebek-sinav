"""Öğrenci listesi Excel/pano okuma: başlık tespiti, esnek (fuzzy) sütun eşleme,
satır ayrıştırma + normalize.

DD `excel_veli.py` (OYS kökenli) kalıbından SADELEŞTİRİLDİ: TCKN, veli
(anne/baba ad-telefon, Veli Kim), doğum tarihi ve cinsiyet sütunları KALDIRILDI —
kelebek bu verileri toplamaz (tasarım §5). e-Okul sınıf/okul listesi ihracının
ya da uygulama şablonunun şu üçlüsü yeter: Sınıf/Şube · Okul No · Ad-Soyad.

Sınıf ayrıştırması okul türünden gelen seviye kümesiyle parametriktir; küme
`parse_rows`'a dışarıdan verilir (parser saf kalır, DB'siz test edilir).
DB eşleştirme/yazma `services/imports.py`'dadır.
"""

from __future__ import annotations

import re
from collections.abc import Collection
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from apps.okul import normalize

# Türkçe karakter → ASCII (başlık eşlemesi için; küçük harfe indirgenir).
_TR_MAP = str.maketrans(
    {
        "ş": "s",
        "Ş": "s",
        "ı": "i",
        "I": "i",
        "İ": "i",
        "ğ": "g",
        "Ğ": "g",
        "ü": "u",
        "Ü": "u",
        "ö": "o",
        "Ö": "o",
        "ç": "c",
        "Ç": "c",
    }
)
_NON_ALNUM = re.compile(r"[^a-z0-9]+")


def normalize_header(text: Any) -> str:
    """Başlık hücresini eşleme için normalleştirir ('Adı Soyadı' → 'adi soyadi')."""
    if text is None:
        return ""
    s = str(text).translate(_TR_MAP).lower()
    s = _NON_ALNUM.sub(" ", s)
    return s.strip()


# Mantıksal alan → normalize edilmiş başlık anahtar kelimeleri (ilk eşleşen kazanır).
COLUMN_SYNONYMS: dict[str, list[str]] = {
    "class": ["sinif sube", "sinif/sube", "sinif", "sube"],
    "number": ["okul no", "okul numarasi", "ogrenci no", "ogrenci numarasi", "numara", "numa"],
    "student_name": [
        "ogrenci adi soyadi",
        "adi soyadi",
        "ad soyad",
        "adsoyad",
        "ad ve soyad",
        "isim",
    ],
    # SIRA KRİTİK: 'soyadi' 'adi'yi içerir — soyad, addan ÖNCE eşlenmeli
    # (excel_personel'deki aynı ders), yoksa 'Öğrenci Soyadı' student_first'e düşer.
    "student_last": ["ogrenci soyadi", "soyadi"],
    "student_first": ["ogrenci adi", "adi"],
}

# Bu sütunlar olmadan içe aktarma yapılamaz (ParserError).
CRITICAL_FIELDS = ("class", "number")


class ParserError(Exception):
    """Girdi ayrıştırılamadığında fırlatılır (kritik sütun eksik vb.)."""


@dataclass
class ColumnMapping:
    """Tespit edilen başlık satırı ve alan → kolon indeksi eşlemesi."""

    header_row: int
    fields: dict[str, int] = field(default_factory=dict)
    matched_headers: dict[str, str] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)

    @property
    def missing_critical(self) -> list[str]:
        missing = [f for f in CRITICAL_FIELDS if f not in self.fields]
        has_combined_name = "student_name" in self.fields
        has_split_name = "student_first" in self.fields and "student_last" in self.fields
        if not has_combined_name and not has_split_name:
            missing.append("student_name veya student_first+student_last")
        return missing

    @property
    def is_usable(self) -> bool:
        return not self.missing_critical


@dataclass
class ParsedRow:
    """Bir veri satırının normalize edilmiş hâli (DB eşleştirmesi öncesi)."""

    row_number: int  # 1-tabanlı (Excel satır no'su)
    raw_class: str = ""
    class_level: int | None = None
    class_section: str = ""
    student_number: str = ""
    raw_student_name: str = ""
    student_first: str = ""
    student_last: str = ""


def _match_field_for_header(norm: str) -> str | None:
    """Normalize başlığı bir mantıksal alana eşler (ilk eşleşen kazanır)."""
    for fieldname, keywords in COLUMN_SYNONYMS.items():
        for kw in keywords:
            if kw == norm or kw in norm:
                return fieldname
    return None


def _map_header_row(cells: list[Any]) -> tuple[dict[str, int], dict[str, str], list[str]]:
    fields: dict[str, int] = {}
    matched: dict[str, str] = {}
    warnings: list[str] = []
    for idx, cell in enumerate(cells):
        norm = normalize_header(cell)
        if not norm:
            continue
        fieldname = _match_field_for_header(norm)
        if fieldname is None:
            continue
        if fieldname in fields:
            warnings.append(
                f"'{cell}' sütunu '{fieldname}' için yinelenen eşleşme; "
                f"ilk eşleşen '{matched[fieldname]}' kullanıldı."
            )
            continue
        fields[fieldname] = idx
        matched[fieldname] = str(cell).strip()
    return fields, matched, warnings


def detect_columns(rows: list[list[Any]], scan_limit: int = 10) -> ColumnMapping:
    """İlk satırlar içinde en iyi başlık satırını bulur ve sütunları eşler."""
    best: ColumnMapping | None = None
    for r in range(min(scan_limit, len(rows))):
        cells = rows[r]
        if all(c is None or str(c).strip() == "" for c in cells):
            continue
        fields, matched, warnings = _map_header_row(cells)
        if not fields:
            continue
        candidate = ColumnMapping(
            header_row=r, fields=fields, matched_headers=matched, warnings=warnings
        )
        if best is None or (candidate.is_usable, len(candidate.fields)) > (
            best.is_usable,
            len(best.fields),
        ):
            best = candidate
        if best.is_usable and best.header_row == r:
            break
    return best or ColumnMapping(header_row=0)


def _cell(cells: list[Any], idx: int | None) -> Any:
    if idx is None or idx >= len(cells):
        return None
    return cells[idx]


def _str(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    # Excel sayıyı float metnine çevirebilir ('2612.0' → '2612').
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def parse_rows(
    rows: list[list[Any]],
    mapping: ColumnMapping,
    *,
    valid_levels: Collection[int] = (9, 10, 11, 12),
) -> list[ParsedRow]:
    """Başlık satırından sonraki veri satırlarını çözümler (boş satırlar atlanır).

    `valid_levels` okul türünden gelir (`SchoolConfig.grade_levels`) — parser
    saf kalsın diye parametredir; servis katmanı doldurur.
    """
    f = mapping.fields
    parsed: list[ParsedRow] = []
    for r in range(mapping.header_row + 1, len(rows)):
        cells = rows[r]
        if all(c is None or str(c).strip() == "" for c in cells):
            continue

        raw_class = _str(_cell(cells, f.get("class")))
        class_parsed = normalize.normalize_class_section(raw_class, valid_levels=valid_levels)
        raw_name = _str(_cell(cells, f.get("student_name")))
        if raw_name:
            first, last = normalize.split_full_name(raw_name)
        else:
            first = _str(_cell(cells, f.get("student_first")))
            last = _str(_cell(cells, f.get("student_last")))
            raw_name = f"{first} {last}".strip()

        parsed.append(
            ParsedRow(
                row_number=r + 1,
                raw_class=raw_class,
                class_level=class_parsed[0] if class_parsed else None,
                class_section=class_parsed[1] if class_parsed else "",
                student_number=_str(_cell(cells, f.get("number"))),
                raw_student_name=raw_name,
                student_first=first,
                student_last=last,
            )
        )
    return parsed


def read_sheet(file_bytes: bytes) -> list[list[Any]]:
    """Excel baytlarını satır-listesine çevirir (etkin sayfa, salt-okunur)."""
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
