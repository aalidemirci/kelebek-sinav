"""Sınav evrak motoru (T8 — yol haritası §9, rapor tasarım sistemi madde 7).

Tüm çıktılar TEK ortak şablondan türer: `templates/sinav/reports/
base.html` — DejaVu Sans (tam Türkçe; kullanıcı kararı: Roboto eklenmedi,
mevcut OYS evrakıyla tipografik tutarlılık), okul + oturum üst bandı, altbilgi
"üretim zamanı + Sayfa x/y", A4 ve gri tonlamalı ofis yazıcısı dostu.

Evrak seti (30.08.2026 sadeleştirmesi — kullanıcı kararı):
- R1  **Salon Sınav Evrakı** (BİRLEŞİK, salon başına 2 yaprak): oturma planı
      krokisi + gözetmen kontrol listesi + evrak sayımı + teslim zinciri
      (yaprak 1) ve yoklama/imza listesi (yaprak 2). Eski R1+R2+R7+R9 yerine
      geçer; çift yüz basıldığında salon başına TEK yaprak düşer.
- R4  Şube Sınav Duyurusu — öğrenci → salon + koltuk; sınıf panosuna asılır.
- R5  Toplu Dağıtım Çizelgesi — Excel (openpyxl), idare çalışma kopyası.
- R6  Gözetmen Görevlendirme ve Tebliğ-Tebellüğ Belgesi.
- R7  **Sınav İhlal ve Kopya Tutanağı** — salon başına bir boş form; birleşik
      salon evrakının tek istisnası (olay tutanağı ayrı yaprak olmak zorunda).
- R8  Dağıtım Doğrulama Raporu — seed + kısıt metrikleri, idare nüshası.
- Ayrıca: oturumdan bağımsız boş salon yerleşim planı (`room_layout.html`).

KALDIRILANLAR: R2 (salon yoklama — R1'e girdi), R2k (şube yoklama — duyuru ve
salon yoklaması ikisini de karşılıyordu), R3 (kapı listesi — kroki + duyuru
zaten söylüyor), R9 (teslim tutanağı — teslim zinciri R1 yaprak 1'e girdi).

KROKİ GEOMETRİ KURALI: çizim GRID kimliğinden — (desk_row, desk_col, slot);
`layout.Seat.x/y` ASLA kullanılmaz (komşu sıra koordinatları çakışabilir —
Tur 223 tuzağı). Bu modül saf veriyle çalışır; DB erişimi services.py'dadır
(booklet.py deseni).

TAŞMA KURALI (kullanıcı kararı): bir derslikte 40 öğrenci sığmalı, fazlası
KONTROLSÜZ taşmamalı. İki mekanizma: (a) `kroki_metrics` krokiyi kendisine
ayrılan kutuya sığdırır — hücre yüksekliği ve punto salonun satır/sütun
sayısından hesaplanır; (b) `list_row_metrics` yoklama/duyuru satırının punto
ve dolgusunu SAYFA BÜTÇESİNDEN türetir; şablon başlık yinelemesi + satır
bölünmezliği listeyi düzgün akıtır. Sayfa sayısı garantileri testte sabittir.
"""

from __future__ import annotations

import io
from collections.abc import Callable
from dataclasses import dataclass

from apps.okul import normalize as okul_normalize
from apps.sinav import layout
from apps.sinav.models import FurnitureKind, SeatStatus

#: Rapor kodu → (başlık, dosya adı kökü). Kodlar sadeleştirme sonrası da
#: KORUNDU (r1/r4/r5/r6/r7/r8) — uç adresleri ve arşivdeki dosya adları
#: kırılmasın; r2/r2k/r3/r9 kaldırıldı (modül açıklamasına bakınız).
REPORT_TITLES: dict[str, tuple[str, str]] = {
    "r1": ("SALON SINAV EVRAKI", "r1_salon_sinav_evraki"),
    "r4": ("ŞUBE SINAV DUYURUSU", "r4_sube_duyurusu"),
    "r5": ("TOPLU DAĞITIM ÇİZELGESİ", "r5_dagitim_cizelgesi"),
    "r6": ("GÖZETMEN GÖREVLENDİRME VE TEBLİĞ-TEBELLÜĞ BELGESİ", "r6_gozetmen_gorevlendirme"),
    "r7": ("SINAV İHLAL VE KOPYA TUTANAĞI", "r7_ihlal_tutanagi"),
    "r8": ("DAĞITIM DOĞRULAMA RAPORU", "r8_dogrulama_raporu"),
}

_FURNITURE_LABELS: dict[str, str] = {
    FurnitureKind.DOOR: "KAPI",
    FurnitureKind.BLACKBOARD: "YAZI TAHTASI",
    FurnitureKind.SMART_BOARD: "AKILLI TAHTA",
    FurnitureKind.TEACHER_DESK: "ÖĞRETMEN MASASI",
}

_STATUS_LABELS: dict[str, str] = {
    SeatStatus.NORMAL: "",
    SeatStatus.PINNED: "Sabit",
    SeatStatus.MANUAL: "Elle taşındı",
}


@dataclass(frozen=True)
class ReportHeader:
    """Ortak üst bant (madde 7): okul + oturum bilgisi + üretim zamanı."""

    school_name: str
    year_label: str
    semester_label: str
    exam_name: str
    exam_date: str  # gg.aa.yyyy
    start_time: str  # SS:DD
    generated_at: str  # gg.aa.yyyy SS:DD


@dataclass(frozen=True)
class SeatRow:
    """Tek yerleşim satırı — SeatAssignment snapshot'ının saf izdüşümü."""

    full_name: str
    student_number: str
    class_label: str
    room_name: str
    seat_no: int
    desk_row: int
    desk_col: int
    slot: int
    course_name: str
    status: str  # SeatStatus değeri


@dataclass(frozen=True)
class RoomSheet:
    """Bir salonun kroki girdisi (plan + o salonun satırları)."""

    room_name: str
    block: str
    plan: layout.LayoutPlan
    numbering_scheme: str
    rows: tuple[SeatRow, ...]


def student_number_sort_key(number: str) -> tuple[int, int | str]:
    """Okul no sıralaması — sayısal numaralar önce ve değerce artan."""
    return (0, int(number)) if number.isdigit() else (1, number)


def room_name_sort_key(name: str) -> tuple[tuple[int, int], ...]:
    """Salon adı sıralaması — TÜRK ALFABESİNE göre ('10/I Dersliği' < '10/İ Dersliği').

    Şube derslikleri `services.section_room_name` ile adlandırılır ve şube harfi
    artık ASCII'ye katlanmadığı için ('10/İ Dersliği') ham `str` karşılaştırması
    Ç/Ğ/İ/Ö/Ş/Ü'yü 'Z'den sonraya atıyordu: 10/I ile 10/İ dersliğinin sayfaları
    basılı evrakın iki ucuna düşüyordu.

    Rakam öbeklerini SAYISAL sıralamaz ('10/A' < '9/A' davranışı korunur) —
    salon adı kelebek düzende serbest metindir ("A-101", "Çok Amaçlı Salon") ve
    karışık tipli bir anahtar sıralamada çöker. Doğal sıralama ayrı bir karardır
    ve TÜM salon yüzeylerinde birlikte yapılmalıdır.
    """
    return okul_normalize.tr_sort_key(name)


def class_label_sort_key(label: str) -> tuple[int, int, tuple[tuple[int, int], ...]]:
    """Şube sıralaması seviye-sayısal, şube harfi TÜRK ALFABESİNE göre.

    9/A < 9/B < 10/B (seviye alfabetik DEĞİL, sayısal) ve 10/C < 10/Ç < 10/D,
    10/I < 10/İ < 10/J. Şube harfi ASCII'ye katlanmadığı için (`normalize.
    tr_upper`) kod noktası sıralaması Ç/Ğ/İ/Ö/Ş/Ü'yü 'Z'den sonraya atardı.
    """
    head, _, section = label.partition("/")
    head = head.strip()
    if head.isdigit():
        return (0, int(head), okul_normalize.tr_sort_key(section))
    return (1, 0, okul_normalize.tr_sort_key(label))


# ---------------------------------------------------------------------------
# Kroki — geometri + ÖLÇÜ (kontrollü taşma)
# ---------------------------------------------------------------------------
# BİRİM UYARISI: WeasyPrint'in iç birimi CSS px'tir (1 pt = 4/3 px) ve düzen
# ölçümleri px cinsinden çıkar. Yükseklik bütçeleri bu yüzden px tutulur ve
# CSS'e px olarak basılır — "hesapladığım = çıkan" kalsın diye. Punto (font)
# baskı alışkanlığı gereği pt kalır.
#
# ÖLÇÜLMÜŞ TUZAK: tablo hücresine `height` vermek satırı KISALTMAZ, UZATIR
# (WeasyPrint hücre yüksekliğini içerik yüksekliğinin üzerine ekler). Liste
# satır ölçüsü bu yüzden PUNTO + DOLGU ile ayarlanır, `height` ile değil.

#: A4 dikey yazım alanı: (297 - 11 - 15) mm × 96/25.4 = 1024 px.
_BUDGET_PX = 1024.0
#: A4 dikey yazım genişliği: (210 - 2×13) mm × 96/25.4 = 695 px.
_CONTENT_WIDTH_PX = 695.0
#: Bir karakterin punto başına genişliği (px): DejaVu Sans ortalama ~0,58 em,
#: em = punto × 4/3 px. Hem kroki hem liste punto kapaklarında kullanılır.
_NAME_CHAR_PX_PER_PT = 0.58 * 4.0 / 3.0
#: Kroki satırları arası `border-spacing` + hücre çerçevesi (ölçüldü, px/satır).
_KROKI_ROW_OVERHEAD_PX = 12.5
#: Bu yüksekliğin altındaki hücreye no + ad + meta üçlüsü sığmaz (meta düşer).
_KROKI_META_MIN_CELL_PX = 30.0
#: Hücre dolgusu + çerçeve payı (px).
_KROKI_CELL_CHROME_PX = 5.0
#: Ad satırı başına hedeflenen karakter — iki kelimelik ad iki satıra sığsın.
_KROKI_LINE_CHARS = 15.0
#: Bir punto başına düşen hücre yüksekliği (px): no + 3 ad satırı (+ meta).
#: no = ad + 0,7 pt · meta = ad - 0,4 pt · satır aralığı ad 1,12 / öteki 1,2.
_KROKI_LINES_WITH_META = 7.68
_KROKI_LINES_COMPACT = 6.08

#: R1 yaprak 1'inde kroki TABLOSUNA ayrılan yükseklik. Yaprağın öteki bölümleri
#: (künye + kontrol listesi + sayım + teslim zinciri + dayanak) ölçülerek
#: ~660 px tuttuğundan krokiye bu kadar kalır.
KROKI_BOX_R1_PX = 340.0
#: Boş salon planında sayfanın neredeyse tamamı krokinindir.
KROKI_BOX_LAYOUT_PX = 760.0


def kroki_metrics(
    rows: int, cols: int, max_seats: int, *, box_height_px: float, with_names: bool
) -> dict[str, object]:
    """Kroki hücre yüksekliği (px) + puntolarını (pt) geometriden hesaplar.

    Amaç TAŞMAYI ÖNLEMEK: hücre yüksekliği `box_height_px` kutusuna bölünür,
    punto hem bu yükseklikten hem de bir koltuğa düşen GENİŞLİKTEN sınırlanır
    (dar sütunda ad taşar). Hücre bir blok kutuya (`.seat-box`) verilir —
    tablo hücresinin `height`i asgarî davranır, blok kutununki bağlayıcıdır.
    Taban/tavan sınırına dayanan çok sıralı salonda kroki kutudan taşabilir;
    taşma KONTROLLÜDÜR (kırpma yok, sonraki bölümler aşağı kayar).

    Değerler METİN döner: TR locale `6.8`'i `6,8` basar ve CSS'te sessizce
    yutulur (F25/T244 tuzağı) — biçimleme burada, nokta ayraçla yapılır.
    """
    rows = max(1, rows)
    cols = max(1, cols)
    seats = max(1, max_seats)

    raw_cell = (box_height_px - rows * _KROKI_ROW_OVERHEAD_PX) / rows
    floor, ceiling = (22.0, 64.0) if with_names else (26.0, 100.0)
    cell = min(max(raw_cell, floor), ceiling)

    seat_width = (_CONTENT_WIDTH_PX / cols) / seats
    if with_names:
        # Punto İKİ kapaktan geçer:
        #  (1) GENİŞLİK — ad satır başına ~15 karakter alabilmeli ki iki
        #      kelimelik bir ad ("ZEYNEP GÜLŞAH" + "KARAOĞLU") iki satıra sığsın;
        #  (2) YÜKSEKLİK — hücre no + ÜÇ ad satırı + meta taşıyabilmeli. Üçüncü
        #      satır payı bilinçli: uzun adlar sarınca meta satırı kırpılıyordu
        #      (kutunun `overflow:hidden`i yarım satır bırakıyordu).
        by_width = (seat_width - _KROKI_CELL_CHROME_PX) / (_KROKI_LINE_CHARS * _NAME_CHAR_PX_PER_PT)
        lines = _KROKI_LINES_WITH_META if cell >= _KROKI_META_MIN_CELL_PX else _KROKI_LINES_COMPACT
        by_height = (cell - _KROKI_CELL_CHROME_PX) / lines
        name = min(max(min(by_height, by_width), 4.6), 7.8)
        no = min(max(name + 0.7, 5.2), 8.6)
        meta = min(max(name - 0.4, 4.2), 7.4)
    else:
        # Boş planda tek içerik koltuk numarasıdır — olabildiğince büyük basılır
        # ve hücrede DİKEY ORTALANIR (`line_height` = hücre yüksekliği).
        name = meta = 0.0
        no = min(max(min(cell * 0.34, seat_width * 0.22), 8.0), 22.0)

    return {
        "cell_height": f"{cell:.1f}",  # px
        "no_font": f"{no:.1f}",  # pt
        "name_font": f"{name:.1f}",  # pt
        "meta_font": f"{meta:.1f}",  # pt
        # Çok sıralı salonda hücre no + ad + meta üçlüsünü taşımaz: meta satırı
        # (okul no · şube) DÜŞER — bilgi yaprak 2'deki yoklama listesinde
        # zaten var; kırpmak yerine kasıtlı sadeleşme (kontrollü taşma).
        "compact": bool(with_names and cell < _KROKI_META_MIN_CELL_PX),
    }


def build_room_kroki(
    sheet: RoomSheet,
    *,
    box_height_px: float = KROKI_BOX_R1_PX,
    with_names: bool = True,
) -> dict[str, object]:
    """Salon krokisi şablon bağlamı: rows×cols hücre matrisi + ölçü sözlüğü.

    Hücre türleri: desk (koltuk kutuları slot sırasında), furniture, empty.
    Devre dışı sıra "KULLANIM DIŞI" olarak çizilir (fiziken salonda durur);
    öğrencisiz aktif koltuk "BOŞ" görünür.
    """
    plan = sheet.plan
    by_seat_key: dict[tuple[int, int, int], SeatRow] = {
        (r.desk_row, r.desk_col, r.slot): r for r in sheet.rows
    }
    seat_no_by_key: dict[tuple[int, int, int], int] = {
        (s.desk_row, s.desk_col, s.slot): s.seat_no
        for s in layout.numbered_seats(plan, sheet.numbering_scheme)
    }
    desk_by_cell = {(d.row, d.col): d for d in plan.desks}
    furniture_by_cell = {(f.row, f.col): f for f in plan.furniture}

    grid: list[list[dict[str, object]]] = []
    for row in range(plan.rows):
        cells: list[dict[str, object]] = []
        for col in range(plan.cols):
            desk = desk_by_cell.get((row, col))
            furn = furniture_by_cell.get((row, col))
            if desk is not None and not desk.disabled:
                seats: list[dict[str, object]] = []
                for slot in range(desk.seat_count):
                    key = (row, col, slot)
                    assigned = by_seat_key.get(key)
                    seats.append(
                        {
                            "seat_no": seat_no_by_key.get(key),
                            "full_name": assigned.full_name if assigned else "",
                            "student_number": assigned.student_number if assigned else "",
                            "class_label": assigned.class_label if assigned else "",
                            "empty": assigned is None,
                        }
                    )
                cells.append({"kind": "desk", "seats": seats})
            elif desk is not None:
                cells.append({"kind": "disabled_desk"})
            elif furn is not None:
                cells.append({"kind": "furniture", "label": _FURNITURE_LABELS[furn.kind]})
            else:
                cells.append({"kind": "empty"})
        grid.append(cells)

    return {
        "room_name": sheet.room_name,
        "block": sheet.block,
        "grid": grid,
        "col_width_pct": round(100.0 / plan.cols, 4),
        "student_count": len(sheet.rows),
        "capacity": plan.capacity,
        "metrics": kroki_metrics(
            plan.rows,
            plan.cols,
            max((d.seat_count for d in plan.desks), default=1),
            box_height_px=box_height_px,
            with_names=with_names,
        ),
    }


# ---------------------------------------------------------------------------
# Liste satır ölçüsü — 40 öğrenci tek sayfaya sığsın (kullanıcı kuralı)
# ---------------------------------------------------------------------------
#: Listenin DIŞINDA kalan sabit yükseklikler (px) — WeasyPrint kutu ağacından
#: ÖLÇÜLDÜ, tahmin değil; `test_reports.py` sayfa sayısıyla sabitler:
#: R1 yaprak 2 = üst bant + bölüm barı + tablo başlığı + imza bloğu + boşluklar.
_ATT_FIXED_PX = 225.0
#: R4 = üst bant + salon dağılımı özeti + bölüm barı + tablo başlığı + kurallar.
_ANN_FIXED_PX = 262.0

#: Ad sütununun sayfa genişliğine oranı (şablondaki sütun yüzdelerinin artığı).
#: R1 yoklama: Sıra 5 + Koltuk 8 + No 9 + Şube 8 + Yok 7 + İmza 27/17
#: (+ Ders 16) → ada 36 % (tek ders) veya 30 % (karışık salon).
#: Şablonda `table-layout: fixed` olduğu için bu oranlar BİREBİR uygulanır.
_ATT_NAME_RATIO, _ATT_NAME_RATIO_MIXED = 0.36, 0.30
#: R4 duyuru: Okul No 10 + Salon 22 + Koltuk 13 (+ Ders 22) → ada 55 % / 33 %.
_ANN_NAME_RATIO, _ANN_NAME_RATIO_MIXED = 0.55, 0.33

#: Satır yüksekliği modeli (ÖLÇÜLDÜ, DejaVu + line-height 1.05, px cinsinden):
#:     satır ≈ 1.4 × punto(pt) + 2.667 × dolgu(pt) + 0.667
_ROW_FONT_COEF = 1.4
_ROW_PAD_COEF = 2.667
_ROW_BORDER_PX = 0.667
#: Gerçek tabloda (kutu sütunu, çok sütunlu hizalama) kalan sabit fark — model
#: ile ölçüm arasındaki artık; kalibrasyonla bulundu.
_ROW_EXTRA_PX = 0.7

#: Punto sınırları: taban okunaklılık, tavan gereksiz büyümeyi önler.
_ROW_FONT_MIN_PT, _ROW_FONT_MAX_PT = 7.0, 10.5
#: Dolgu tavanı — satır seyreltmesi bir yere kadar (px değil, pt).
_ROW_PAD_MAX_PT = 4.0

#: Ad sütunu TEK SATIRA sığmalı: sarma satırı iki katına çıkarır ve sayfa
#: garantisini bozar. Uzun bir Türkçe ad-soyad ~28 karakterdir.
_NAME_MAX_CHARS = 28.0
#: Hücre yatay dolgusu + çerçeve payı (px).
#: 31.08.2026: `.att`/`.ann` tabloları `box-sizing: border-box` oldu (sütun
#: yüzdeleri content-box'ta dolguyu dışarı ekleyip tabloyu sayfadan taşırıyordu
#: — R1'de 91pt, R4'te 57pt ÖLÇÜLDÜ). Artık yatay dolgu (5pt + 5pt = 13,3px)
#: sütunun İÇİNDEN çıkıyor; ad genişliği payı o kadar büyütüldü. Küçülen payla
#: punto da küçülür — yön GÜVENLİ taraftadır (satır kısalır).
_NAME_CELL_CHROME_PX = 14.0 + 10.0 * 4.0 / 3.0


def list_row_metrics(count: int, *, fixed_px: float, name_col_ratio: float) -> dict[str, str]:
    """Liste satırının PUNTO ve DOLGU değerlerini sayfa bütçesinden hesaplar.

    Kademeli sınıf yerine sürekli değer: hedef satır = (bütçe - sabitler) / n.
    Hedefe önce puntoyla, artan boşluğa dolguyla ulaşılır. `height` KULLANILMAZ
    — WeasyPrint'te hücre yüksekliği satırı uzatır (ölçüldü).

    Punto ayrıca AD SÜTUNU GENİŞLİĞİNDEN sınırlanır (`name_col_ratio`, sayfa
    genişliğine oran): sarmayan ad = öngörülebilir satır yüksekliği. Ders
    sütunu açıldığında ad sütunu daralır ve punto kendiliğinden küçülür.

    Böylece 40 öğrenci tek sayfaya SIĞAR (garanti testle sabitlenir); punto
    tabanına dayanan çok kalabalık salonda liste bölünmeden akar (başlık
    yinelenir, imza bloğu bütün hâlde kayar) — kontrolsüz taşma olmaz.

    Değerler METİN döner (TR locale ondalığı virgülle basar — CSS yutar).
    """
    rows = max(1, count)
    target = (_BUDGET_PX - fixed_px) / rows
    by_height = (target - _ROW_EXTRA_PX - _ROW_BORDER_PX) / _ROW_FONT_COEF
    by_width = (_CONTENT_WIDTH_PX * name_col_ratio - _NAME_CELL_CHROME_PX) / (
        _NAME_MAX_CHARS * _NAME_CHAR_PX_PER_PT
    )
    font = min(max(min(by_height, by_width), _ROW_FONT_MIN_PT), _ROW_FONT_MAX_PT)
    pad = (target - _ROW_EXTRA_PX - _ROW_FONT_COEF * font - _ROW_BORDER_PX) / _ROW_PAD_COEF
    pad = min(max(pad, 0.0), _ROW_PAD_MAX_PT)
    # İşaret kutusu satır kutusunu YÜKSELTMEMELİ: satır içi blok, metin
    # satırından yüksekse satır kutusunu büyütür (ölçüldü: 8 pt kutu satıra
    # ~3,3 px ekliyordu). Punto çizgisinin altında kalacak boyut seçilir.
    box = min(7.5, max(4.5, font * 0.80))
    return {
        "font_size": f"{font:.2f}",
        "padding": f"{pad:.2f}",
        "box_size": f"{box:.2f}",
    }


def _course_breakdown(rows: list[SeatRow] | tuple[SeatRow, ...]) -> list[dict[str, object]]:
    """Ders → kayıtlı sayısı (deste sayımı ve künye özeti için)."""
    counts: dict[str, int] = {}
    for row in rows:
        counts[row.course_name or "—"] = counts.get(row.course_name or "—", 0) + 1
    return [{"course_name": name, "count": count} for name, count in sorted(counts.items())]


def _course_summary(courses: list[dict[str, object]]) -> str:
    """Künye satırı: "Coğrafya (20) · Matematik (18)"."""
    return " · ".join(f"{c['course_name']} ({c['count']})" for c in courses) or "—"


# ---------------------------------------------------------------------------
# R1 — birleşik salon sınav evrakı (kroki + gözetmen işlemleri + yoklama)
# ---------------------------------------------------------------------------
def build_room_documents(
    sheets: list[RoomSheet], *, proctor_names: dict[str, str] | None = None
) -> list[dict[str, object]]:
    """R1: salon başına İKİ yapraklık tek belge bağlamı.

    Yaprak 1 künye + kroki + gözetmen kontrol listesi + evrak sayımı + teslim
    zinciri; yaprak 2 koltuk sırasında yoklama/imza listesi. `proctor_names`
    (salon adı → görevli) doluysa gözetmen adı BASILI gelir; boşsa alan elle
    doldurulur (gözetmen modülü kapalı — K2).

    `sheets` sırası çağıranın verdiği sıradır (services `_room_sheets` salon
    adını Türk alfabesine göre dizer) — basılı evrağın sayfa sırası budur.
    """
    names = proctor_names or {}
    documents: list[dict[str, object]] = []
    for sheet in sheets:
        ordered = sorted(sheet.rows, key=lambda r: r.seat_no)
        courses = _course_breakdown(ordered)
        documents.append(
            {
                "room_name": sheet.room_name,
                "block": sheet.block,
                "kroki": build_room_kroki(sheet, box_height_px=KROKI_BOX_R1_PX),
                "rows": ordered,
                "registered": len(ordered),
                "capacity": sheet.plan.capacity,
                "courses": courses,
                "course_summary": _course_summary(courses),
                # Ders sütunu yalnız KARIŞIK salonda anlamlı — tek derste
                # sütun yerine imza alanı genişler.
                "show_course": len(courses) > 1,
                "row": list_row_metrics(
                    len(ordered),
                    fixed_px=_ATT_FIXED_PX,
                    name_col_ratio=_ATT_NAME_RATIO_MIXED if len(courses) > 1 else _ATT_NAME_RATIO,
                ),
                "proctor_name": names.get(sheet.room_name, ""),
            }
        )
    return documents


# ---------------------------------------------------------------------------
# R4 — şube duyurusu · R7 — ihlal tutanağı
# ---------------------------------------------------------------------------
def build_announcements(rows: list[SeatRow]) -> list[dict[str, object]]:
    """R4: şube başına sayfa, okul no sırasında — öğrenci → salon + koltuk.

    `room_summary` duyurunun en çok okunan satırıdır ("şubem nereye dağıldı"):
    salon adları Türk alfabesi sırasında, yanlarında öğrenci sayısıyla.
    """
    sheets: list[dict[str, object]] = []
    for class_label, group in _grouped(
        rows, key=lambda r: r.class_label, sort_key=class_label_sort_key
    ).items():
        ordered = sorted(group, key=lambda r: student_number_sort_key(r.student_number))
        mixed = len({row.course_name for row in ordered}) > 1
        room_counts: dict[str, int] = {}
        for row in ordered:
            room_counts[row.room_name] = room_counts.get(row.room_name, 0) + 1
        sheets.append(
            {
                "class_label": class_label,
                "rows": ordered,
                "room_summary": " · ".join(
                    f"{name} ({count})"
                    for name, count in sorted(
                        room_counts.items(), key=lambda kv: room_name_sort_key(kv[0])
                    )
                ),
                "row": list_row_metrics(
                    len(ordered),
                    fixed_px=_ANN_FIXED_PX,
                    name_col_ratio=_ANN_NAME_RATIO_MIXED if mixed else _ANN_NAME_RATIO,
                ),
                "show_course": mixed,
            }
        )
    return sheets


def build_tutanak_sheets(
    rows: list[SeatRow], *, proctor_names: dict[str, str] | None = None
) -> list[dict[str, object]]:
    """R7: salon başına BİR boş ihlal/kopya tutanağı (salon zarfına konur).

    Öğrenci alanları BOŞTUR — olay önceden bilinemez ve kişisel veri basılmaz;
    yalnız salon/ders künyesi ve (varsa) gözetmen adı basılı gelir.
    """
    names = proctor_names or {}
    return [
        {
            "room_name": room_name,
            "course_summary": _course_summary(_course_breakdown(group)),
            "proctor_name": names.get(room_name, ""),
        }
        for room_name, group in _grouped(
            rows, key=lambda r: r.room_name, sort_key=room_name_sort_key
        ).items()
    ]


def _grouped(
    rows: list[SeatRow],
    *,
    key: Callable[[SeatRow], str],
    sort_key: Callable[[str], tuple[object, ...]] | None = None,
) -> dict[str, list[SeatRow]]:
    """Anahtar değerine göre gruplar; grup sırası `sort_key` (yoksa alfabetik)."""
    groups: dict[str, list[SeatRow]] = {}
    for row in rows:
        groups.setdefault(key(row), []).append(row)
    if sort_key is None:
        return dict(sorted(groups.items()))
    return dict(sorted(groups.items(), key=lambda kv: sort_key(kv[0])))


# ---------------------------------------------------------------------------
# R6 — gözetmen görevlendirme / tebliğ-tebellüğ (T9b)
# ---------------------------------------------------------------------------
@dataclass(frozen=True)
class ProctorRow:
    """Tek görevlendirme satırı — ProctorAssignment snapshot'ının saf izdüşümü."""

    teacher_name: str
    role: str  # ProctorRole değeri
    role_label: str
    room_name: str  # YEDEK için boş


#: R6 tablo sırası: salonlu görevler salon adına göre; yedekler sonda (Tur 235: CHIEF kalktı).
_PROCTOR_ROLE_ORDER: dict[str, int] = {"PROCTOR": 0, "RESERVE": 1}


def build_assignment_context(rows: list[ProctorRow]) -> dict[str, object]:
    """R6 şablon bağlamı — resmî yazı tablosu + sayım satırları."""
    ordered = sorted(
        rows,
        key=lambda r: (
            r.room_name == "",  # yedekler (salonsuz) sona
            room_name_sort_key(r.room_name),
            _PROCTOR_ROLE_ORDER.get(r.role, 9),
            r.teacher_name,
        ),
    )
    return {
        "rows": [
            {
                "teacher_name": row.teacher_name,
                "role_label": row.role_label,
                "room_name": row.room_name or "—",
            }
            for row in ordered
        ],
        "duty_count": len(ordered),
        "reserve_count": sum(1 for row in ordered if not row.room_name),
        "room_count": sum(1 for row in ordered if row.room_name),
    }


# ---------------------------------------------------------------------------
# R8 — dağıtım doğrulama bağlamı
# ---------------------------------------------------------------------------
def build_validation_context(
    *,
    is_valid: bool,
    hard_violations: list[str],
    first_ring_pairs: int,
    min_distances: dict[str, float],
    proximity_score: float,
    params: dict[str, object],
    group_labels: dict[str, str],
    warnings: list[str],
    cross_section_pairs: int = 0,
    occupancy: list[dict[str, object]] | None = None,
) -> dict[str, object]:
    """R8 şablon bağlamı — metrik anahtarları ders adına çözülmüş halde.

    İhlal metinleri doğrulayıcıdan adsız gelir (okul no bile yok, yalnız
    id/koordinat); burada da kişisel veri eklenmez. K1 (Tur 645):
    `cross_section_pairs` (aynı şube, farklı grup, 1. halka) +
    `occupancy` (salon doluluk tablosu — salon adı + sayı, PII yok).
    """
    return {
        "is_valid": is_valid,
        "hard_violations": hard_violations,
        "violation_count": len(hard_violations),
        "first_ring_pairs": first_ring_pairs,
        "min_distances": [
            {"group_label": group_labels.get(key, key), "distance": dist}
            for key, dist in sorted(min_distances.items())
        ],
        "proximity_score": round(proximity_score, 4),
        "cross_section_pairs": cross_section_pairs,
        "occupancy": list(occupancy or []),
        "params": params,
        "warnings": warnings,
    }


def reports_zip(files: list[tuple[str, bytes]]) -> bytes:
    """Üretilmiş evrakları tek ZIP'te toplar (dosya adı güvenli karaktere indirgenir)."""
    import zipfile

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, content in files:
            safe = "".join(c if c.isalnum() or c in "-_. " else "_" for c in filename)
            zf.writestr(safe or "evrak", content)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Render — PDF (WeasyPrint) + Excel (openpyxl)
# ---------------------------------------------------------------------------
def render_pdf(template_name: str, context: dict[str, object]) -> bytes:
    """Ortak şablonu kullanan rapor sayfasını PDF'e çevirir."""
    from django.template.loader import render_to_string
    from weasyprint import HTML  # tembel import — ağır bağımlılık

    return bytes(HTML(string=render_to_string(template_name, context)).write_pdf())


def build_r5_workbook(header: ReportHeader, rows: list[SeatRow]) -> bytes:
    """R5: tüm oturum tek sayfa Excel — idare çalışma kopyası.

    raporlar modülünün exporter'ı İÇERİ AKTARILMAZ (ADR-0002 modül sınırı);
    çizelge burada doğrudan openpyxl ile kurulur. Gri tonlamalı stil.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    title, _ = REPORT_TITLES["r5"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Dagitim"

    columns = ("Okul No", "Ad Soyad", "Şube", "Ders", "Salon", "Koltuk No", "Durum")
    ws.cell(row=1, column=1, value=f"{header.school_name} — {title}").font = Font(
        bold=True, size=13
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    meta = (
        f"{header.exam_name} · {header.exam_date} {header.start_time} · "
        f"{header.year_label} {header.semester_label} · Üretim: {header.generated_at}"
    )
    ws.cell(row=2, column=1, value=meta)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(columns))

    header_row = 4
    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9D9D9")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ordered = sorted(
        rows, key=lambda r: (room_name_sort_key(r.room_name), r.seat_no)
    )  # salon + oturma sırası — deste/karşılaştırma kopyası
    for offset, row in enumerate(ordered):
        values = (
            row.student_number,
            row.full_name,
            row.class_label,
            row.course_name,
            row.room_name,
            row.seat_no,
            _STATUS_LABELS.get(row.status, row.status),
        )
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=header_row + 1 + offset, column=col_idx, value=value)

    widths = (10, 32, 8, 26, 18, 10, 14)
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = f"A{header_row + 1}"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
