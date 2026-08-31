"""F4 evrak testleri — R1 · R4-R8 + boş plan + tümü-ZIP (WeasyPrint + pypdf).

Kapı (tasarım §12 F4):
- her raporda TR karakter duman testi (DD `test_documents.py` emsali —
  pypdf metin çıkarma; dar kapsama fontta Ğ/Ş/İ sessizce düşer),
- `text-transform` tarama testi (WeasyPrint TR i→I tuzağı — CLAUDE.md §2),
- `|unlocalize` denetimi (TR locale ondalığı virgülle basar; CSS genişliği
  yutulur — OYS F25/T244 bulgusu),
- **SAYFA SAYISI garantisi**: bir derslikte 40 öğrenci sığar ve fazlası
  kontrolsüz taşmaz (kullanıcı kuralı, 30.08.2026 sadeleştirmesi).
Durum kapıları (taslak reddi, arşivden yeniden basım) ve R8 seed sözleşmesi
(CLAUDE.md §3) burada sabitlenir.
"""

from __future__ import annotations

import io
import re
import zipfile
from pathlib import Path
from typing import Any

import pytest
from django.conf import settings
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
from pypdf import PdfReader
from rest_framework.test import APIClient

from apps.okul.models import SchoolConfig
from apps.sinav import layout, reports, services
from apps.sinav.models import ExamSession, ExamSessionRoom
from apps.sinav.tests.oturum_yardim import dagitilmis_oturum, oturum, salon

pytestmark = pytest.mark.django_db

#: DD emsali duman metni — Türkçe glifler kayıpsız çıkmalı.
TURKCE_DUMAN = "ĞÜŞİÖÇ ığüşiöç"
OKUL_ADI = f"{TURKCE_DUMAN} Anadolu Lisesi"

#: PDF üreten oturum raporları (r5 Excel, r6 F7'de).
PDF_CODES = ("r1", "r4", "r7", "r8")

#: 30.08.2026 sadeleştirmesinde KALDIRILAN kodlar — geri sızarsa test kırılır.
KALDIRILAN_CODES = ("r2", "r2k", "r3", "r9")

SESSIONS_URL = "/api/v1/exam-sessions/"
ROOMS_URL = "/api/v1/exam-rooms/"

_TEMPLATES_DIR = Path(settings.BASE_DIR) / "templates"


def _pdf_text(pdf_bytes: bytes) -> str:
    """PDF gövde metni (DD `_pdf_text` deseni)."""
    reader = PdfReader(io.BytesIO(pdf_bytes))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _okul() -> None:
    """Rapor başlığına Türkçe duman metnini okul adı üzerinden enjekte eder."""
    SchoolConfig.objects.create(pk=SchoolConfig.SINGLETON_PK, school_name=OKUL_ADI)


def _evrak_oturumu(**kwargs: Any) -> ExamSession:
    """Okul yapılandırması + dağıtılmış oturum (evrak üretimine hazır)."""
    _okul()
    return dagitilmis_oturum(**kwargs)


# ===========================================================================
# TR karakter duman testi — her raporda (F4 kapısı)
# ===========================================================================


@pytest.mark.parametrize("code", PDF_CODES)
def test_pdf_raporlarda_turkce_duman(code: str) -> None:
    session = _evrak_oturumu()
    rf = services.render_session_report(session, code)

    _title, stem = reports.REPORT_TITLES[code]
    assert rf.filename == f"{stem}_oturum_{session.pk}.pdf"
    assert rf.content_type == "application/pdf"
    assert rf.content.startswith(b"%PDF")

    text = _pdf_text(rf.content)
    # DD deseni: boşluk hariç HARF HARF (extract_text satır sonu ekleyebilir).
    eksik = [harf for harf in TURKCE_DUMAN if harf != " " and harf not in text]
    assert not eksik, f"{code} çıktısında Türkçe glif kaybı: {eksik}"


def test_bos_salon_plani_turkce_duman() -> None:
    _okul()
    s = salon("Şölen İçi Derslik")  # salon adı da Türkçe glif taşır
    rf = services.render_room_layout_pdf(s)

    assert rf.filename == f"salon_yerlesim_plani_{s.pk}.pdf"
    assert rf.content.startswith(b"%PDF")
    text = _pdf_text(rf.content)
    eksik = [harf for harf in TURKCE_DUMAN if harf != " " and harf not in text]
    assert not eksik, f"Boş plan çıktısında Türkçe glif kaybı: {eksik}"
    assert "Şölen İçi Derslik" in text


def test_r5_excel_cizelge_turkce_duman() -> None:
    session = _evrak_oturumu()
    rf = services.render_session_report(session, "r5")

    assert rf.filename == f"r5_dagitim_cizelgesi_oturum_{session.pk}.xlsx"
    assert rf.content_type == ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    wb = load_workbook(io.BytesIO(rf.content))
    ws = wb["Dagitim"]
    assert ws.freeze_panes == "A5"
    baslik = str(ws["A1"].value)
    assert OKUL_ADI in baslik and "TOPLU DAĞITIM ÇİZELGESİ" in baslik
    kolonlar = tuple(ws.cell(row=4, column=c).value for c in range(1, 8))
    assert kolonlar == ("Okul No", "Ad Soyad", "Şube", "Ders", "Salon", "Koltuk No", "Durum")
    # 2 seviye × 3 öğrenci = 6 veri satırı (5. satırdan itibaren).
    assert ws.max_row == 4 + 6


# ===========================================================================
# R8 sözleşmesi — seed basılır (CLAUDE.md §3)
# ===========================================================================


def test_r8_seed_ve_ders_etiketi_basilir() -> None:
    session = _evrak_oturumu(seed=987654)
    text = _pdf_text(services.render_session_report(session, "r8").content)
    assert "987654" in text, "R8 dağıtım seed'ini basmalı (aynı seed → aynı dağıtım kanıtı)"
    # Çakışma grubu anahtarı ham değil, ders adlı etiketle görünmeli.
    assert "Coğrafya" in text


# ===========================================================================
# Tümü-ZIP
# ===========================================================================


def test_zip_tum_evrak() -> None:
    session = _evrak_oturumu()
    rf = services.render_session_reports_zip(session)

    assert rf.filename == f"sinav_evraki_oturum_{session.pk}.zip"
    assert rf.content_type == "application/zip"
    with zipfile.ZipFile(io.BytesIO(rf.content)) as zf:
        adlar = sorted(zf.namelist())
    beklenen = sorted(
        f"{reports.REPORT_TITLES[code][1]}_oturum_{session.pk}"
        + (".xlsx" if code == "r5" else ".pdf")
        for code in services.REPORT_CODES
        if code != "r6"  # görevlendirme modeli F7'de — r6 pakete girmez
    )
    assert adlar == beklenen


# ===========================================================================
# Durum kapıları + hata yolları
# ===========================================================================


def test_taslak_oturumda_evrak_reddedilir() -> None:
    _okul()
    draft = oturum(name="Taslak Oturum")
    with pytest.raises(ValidationError, match="Önce dağıtım"):
        services.render_session_report(draft, "r1")


def test_arsivden_yeniden_basim_acik() -> None:
    session = _evrak_oturumu()
    session = services.approve_session(session, approved_by_name="Örnek MÜDÜR")
    session = services.archive_session(session)
    rf = services.render_session_report(session, "r1")
    assert rf.content.startswith(b"%PDF")
    zip_rf = services.render_session_reports_zip(session)
    assert zip_rf.content_type == "application/zip"


def test_bilinmeyen_rapor_kodu() -> None:
    session = _evrak_oturumu()
    with pytest.raises(ValidationError, match="Bilinmeyen rapor kodu"):
        services.render_session_report(session, "r99")


def test_r6_f4te_uretilmez() -> None:
    kapali = _evrak_oturumu()
    with pytest.raises(ValidationError, match="kapalı"):
        services.render_session_report(kapali, "r6")

    # Gözetmen ayarı açık ama görevlendirme modeli F7'de — guard durumdan
    # hemen sonra çalıştığından yerleşimsiz DAĞITILDI kabuğu yeterli.
    acik = oturum(name="Gözetmenli Oturum", proctors_enabled=True)
    acik.status = kapali.status
    acik.save(update_fields=["status"])
    with pytest.raises(ValidationError, match="Görevlendirme yapılmamış"):
        services.render_session_report(acik, "r6")


def test_salon_filtresi() -> None:
    session = _evrak_oturumu(rooms=2, per_level=6)
    oturum_salonlari = list(
        ExamSessionRoom.objects.filter(session=session).select_related("room").order_by("order")
    )
    ilk = oturum_salonlari[0].room

    text = _pdf_text(services.render_session_report(session, "r1", room_id=ilk.pk).content)
    assert ilk.name in text
    digerleri = [sr.room.name for sr in oturum_salonlari[1:]]
    assert all(ad not in text for ad in digerleri)

    with pytest.raises(ValidationError, match="salon bazlı"):
        services.render_session_report(session, "r4", room_id=ilk.pk)
    # R7 tutanağı da salon bazlıdır (salon zarfına konur).
    services.render_session_report(session, "r7", room_id=ilk.pk)
    with pytest.raises(ValidationError, match="tanımlı değil"):
        services.render_session_report(session, "r1", room_id=999999)


# ===========================================================================
# API uçları
# ===========================================================================


def test_api_rapor_indirme() -> None:
    session = _evrak_oturumu()
    client = APIClient()

    resp = client.get(f"{SESSIONS_URL}{session.pk}/reports/r1/")
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/pdf"
    assert (
        f'filename="r1_salon_sinav_evraki_oturum_{session.pk}.pdf"' in resp["Content-Disposition"]
    )
    assert resp.content.startswith(b"%PDF")

    resp = client.get(f"{SESSIONS_URL}{session.pk}/reports/zip/")
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/zip"

    resp = client.get(f"{SESSIONS_URL}{session.pk}/reports/r99/")
    assert resp.status_code == 404
    assert resp.json()["code"] == "not_found"

    # Taslakta Türkçe 400 (servis kapısı uca yansır).
    draft = oturum(name="Taslak Oturum")
    resp = client.get(f"{SESSIONS_URL}{draft.pk}/reports/r1/")
    assert resp.status_code == 400


def test_api_bos_salon_plani() -> None:
    _okul()
    s = salon("D-101")
    resp = APIClient().get(f"{ROOMS_URL}{s.pk}/layout-pdf/")
    assert resp.status_code == 200
    assert resp["Content-Type"] == "application/pdf"
    assert f'filename="salon_yerlesim_plani_{s.pk}.pdf"' in resp["Content-Disposition"]


# ===========================================================================
# SAYFA SAYISI GARANTİSİ — "bir derslikte 40 öğrenci sığsın, fazlası
# kontrolsüz taşmasın" (kullanıcı kuralı, 30.08.2026 sadeleştirmesi)
# ===========================================================================
#
# Bu blok saf bağlamla (DB'siz) çalışır: kalabalık salon kurmak 40+ öğrenci ve
# büyük yerleşim planı ister; bunu ORM'den kurmak testi yavaşlatır ve garanti
# zaten dizgi/ölçü katmanına (reports.py + şablon) aittir.

#: Uzun ama gerçekçi Türkçe adlar — satır sarması en kötü hâlde sınanır.
_UZUN_ADLAR = (
    "ZEYNEP GÜLŞAH KARAOĞLU",
    "MEHMET ALİ ÇAĞLAYANOĞLU",
    "ABDÜLKADİR ŞAHİNKAYA",
    "ÖZGE NUR BÜYÜKKAYAOĞLU",
)

_BASLIK = reports.ReportHeader(
    school_name=OKUL_ADI,
    year_label="2026-2027",
    semester_label="1. Dönem",
    exam_name="1. Ortak Yazılı Sınav",
    exam_date="16.11.2026",
    start_time="09:00",
    generated_at="30.08.2026 19:45",
)


def _plan(rows: int, cols: int) -> dict[str, Any]:
    """rows×cols ikili sıra ızgarası (demirbaşsız)."""
    return {
        "grid": {"rows": rows, "cols": cols},
        "desks": [{"row": r, "col": c, "type": "DOUBLE"} for r in range(rows) for c in range(cols)],
        "furniture": [],
    }


def _satirlar(n: int, *, dersler: tuple[str, ...]) -> list[reports.SeatRow]:
    return [
        reports.SeatRow(
            full_name=_UZUN_ADLAR[i % len(_UZUN_ADLAR)],
            student_number=str(1200 + i),
            class_label=f"{9 + i % 4}/{'ABCÇ'[i % 4]}",
            room_name="D-201 Dersliği",
            seat_no=i + 1,
            desk_row=i // 8,
            desk_col=(i % 8) // 2,
            slot=i % 2,
            course_name=dersler[i % len(dersler)],
            status="NORMAL",
        )
        for i in range(n)
    ]


def _sayfa_sayisi(pdf: bytes) -> int:
    return len(PdfReader(io.BytesIO(pdf)).pages)


def _r1_pdf(n: int, rows: int, cols: int, dersler: tuple[str, ...]) -> bytes:
    sheet = reports.RoomSheet(
        room_name="D-201 Dersliği",
        block="A Blok · 2. kat",
        plan=layout.validate_layout_plan(_plan(rows, cols)),
        numbering_scheme="S_PATTERN",
        rows=tuple(_satirlar(n, dersler=dersler)),
    )
    return reports.render_pdf(
        "sinav/reports/r1_salon_evraki.html",
        {
            "header": _BASLIK,
            "title": reports.REPORT_TITLES["r1"][0],
            "sheets": reports.build_room_documents([sheet]),
        },
    )


@pytest.mark.parametrize(
    ("ogrenci", "rows", "cols", "ders_sayisi"),
    [
        (40, 5, 4, 2),  # tipik derslik: 5 sıra × 4 ikili masa
        (40, 8, 3, 2),  # derin derslik
        (40, 10, 2, 1),  # dar ve uzun derslik
        (40, 4, 5, 3),  # geniş derslik + karışık ders (Ders sütunu açılır)
        (48, 6, 4, 2),  # kapasite üstü — yine iki yaprak
        (12, 3, 2, 1),  # küçük salon: satırlar seyreler, yine iki yaprak
    ],
)
def test_r1_salon_evraki_iki_yaprak(ogrenci: int, rows: int, cols: int, ders_sayisi: int) -> None:
    """Salon evrakı HER geometride ve 40 öğrenciye kadar TAM İKİ yaprak olmalı.

    Kullanıcı kuralı: bir derslikte 40 öğrenci sığar, fazlası kontrolsüz
    taşmaz. İki yaprak = çift yüz basıldığında salon başına tek kâğıt.
    Kırılırsa bakılacak yer: `reports.KROKI_BOX_R1_PX`, `_ATT_FIXED_PX` ve
    yaprak 1'in sabit bölümleri (şablon yorumundaki sayfa bütçesi).
    """
    dersler = tuple(f"Ders {i}" for i in range(ders_sayisi))
    pdf = _r1_pdf(ogrenci, rows, cols, dersler)
    assert (
        _sayfa_sayisi(pdf) == 2
    ), f"{ogrenci} öğrenci / {rows}x{cols} salonda yaprak sayısı 2 değil — sayfa bütçesi bozuldu."


def test_r1_cok_kalabalik_salonda_satir_kaybi_yok() -> None:
    """40'ı çok aşan salonda liste TAŞAR ama satır KAYBOLMAZ (kontrollü taşma)."""
    pdf = _r1_pdf(90, 6, 4, ("Coğrafya",))
    metin = _pdf_text(pdf)
    assert _sayfa_sayisi(pdf) > 2, "90 öğrencide listenin akması beklenir"
    for numara in ("1200", "1245", "1289"):  # ilk, orta, son
        assert numara in metin, f"{numara} numaralı öğrenci evrakta yok — satır düştü"


def test_r4_sube_duyurusu_tek_yaprak() -> None:
    """Şube duyurusu 40 öğrencide tek sayfa — ders sütunu açıkken de."""
    for dersler in (("Coğrafya",), ("Coğrafya", "Matematik")):
        satirlar = [
            reports.SeatRow(**{**vars(r), "class_label": "9/A", "room_name": f"D-20{i % 3 + 1}"})
            for i, r in enumerate(_satirlar(40, dersler=dersler))
        ]
        pdf = reports.render_pdf(
            "sinav/reports/r4_announcement.html",
            {
                "header": _BASLIK,
                "title": reports.REPORT_TITLES["r4"][0],
                "sheets": reports.build_announcements(satirlar),
            },
        )
        assert _sayfa_sayisi(pdf) == 1, f"{len(dersler)} derslik duyuru tek sayfa değil"


def test_r1_birlesik_evrak_dort_belgenin_isini_tasir() -> None:
    """Birleşik evrak eski R1 + R2 + R7 + R9'un işlerini TEK belgede taşımalı."""
    metin = _pdf_text(_r1_pdf(24, 4, 3, ("Coğrafya",)))
    for beklenen in (
        "OTURMA PLANI",  # eski R1 krokisi
        "YOKLAMA VE İMZA LİSTESİ",  # eski R2
        "GÖZETMEN İŞLEMLERİ",  # kullanıcı talebi
        "SINAV EVRAKI SAYIMI",  # eski R7 deste sayımı
        "EVRAK TESLİM ZİNCİRİ",  # eski R9 teslim tutanağı
    ):
        assert beklenen in metin, f"birleşik evrakta eksik bölüm: {beklenen}"


def test_r7_tutanak_bos_formdur() -> None:
    """İhlal tutanağı salon başına tek yaprak; öğrenci verisi BASILMAZ (KVKK)."""
    pdf = reports.render_pdf(
        "sinav/reports/r7_tutanak.html",
        {
            "header": _BASLIK,
            "title": reports.REPORT_TITLES["r7"][0],
            "sheets": reports.build_tutanak_sheets(_satirlar(40, dersler=("Coğrafya",))),
        },
    )
    assert _sayfa_sayisi(pdf) == 1
    metin = _pdf_text(pdf)
    assert "D-201 Dersliği" in metin  # salon künyesi basılı gelir
    for ad in _UZUN_ADLAR:
        assert ad not in metin, "tutanak boş formdur — öğrenci verisi basılmaz"


def test_kaldirilan_rapor_kodlari_geri_gelmedi() -> None:
    """R2/R2k/R3/R9 kaldırıldı — kod, başlık ve şablon geri sızmamalı."""
    for code in KALDIRILAN_CODES:
        assert code not in services.REPORT_CODES
        assert code not in reports.REPORT_TITLES
    for sablon in ("r1_kroki.html", "r2_attendance.html", "r3_door.html", "r9_handover.html"):
        assert not (_TEMPLATES_DIR / "sinav" / "reports" / sablon).exists()


# ===========================================================================
# Şablon tarama kapıları (F4) — dosya sistemi denetimleri
# ===========================================================================

#: Yorum blokları taramadan düşülür — yasak hatırlatma metinleri serbest,
#: aranan GERÇEK CSS bildirimidir.
_YORUM_KALIBI = re.compile(
    r"\{%\s*comment\s*%\}.*?\{%\s*endcomment\s*%\}|<!--.*?-->|/\*.*?\*/", re.DOTALL
)


def test_sablonlarda_text_transform_yasak() -> None:
    """`text-transform:` bildirimi YASAK — WeasyPrint TR i→I basar (CLAUDE.md §2)."""
    ihlaller: list[str] = []
    for dosya in sorted(_TEMPLATES_DIR.rglob("*")):
        if dosya.suffix not in (".html", ".css"):
            continue
        icerik = _YORUM_KALIBI.sub("", dosya.read_text(encoding="utf-8"))
        if re.search(r"text-transform\s*:", icerik):
            ihlaller.append(str(dosya.relative_to(_TEMPLATES_DIR)))
    assert not ihlaller, f"text-transform bildirimi bulunan şablonlar: {ihlaller}"


def test_unlocalize_denetimi() -> None:
    """CSS'e giren ondalıklı değişkenler `|unlocalize` taşımalı (F25/T244).

    TR locale `25.0`'ı `25,0` basar; `width: 25,0%` sessizce yutulur ve
    kroki tablosu çöker. Kural: `pct`/`percent` adlı her şablon değişkeni
    unlocalize'lı olmalı; bilinen üç kullanım da yerinde sabitlenir.
    """
    for dosya in sorted(_TEMPLATES_DIR.rglob("*.html")):
        icerik = dosya.read_text(encoding="utf-8")
        for degisken in re.findall(r"\{\{[^}]*\}\}", icerik):
            if ("pct" in degisken or "percent" in degisken) and "|unlocalize" not in degisken:
                pytest.fail(f"{dosya.name}: unlocalize'sız ondalık değişken: {degisken}")

    reports_dir = _TEMPLATES_DIR / "sinav" / "reports"
    # Kroki artık TEK parçadan gelir (R1 ve boş plan onu paylaşır).
    assert "col_width_pct|unlocalize" in (reports_dir / "_kroki.html").read_text("utf-8")
    assert "percent|unlocalize" in (reports_dir / "r8_validation.html").read_text("utf-8")


def test_design_css_ayni_kaldi() -> None:
    """`_design.css` içine Django etiketi yazılamaz (kendini include — OYS Tur 238)."""
    icerik = (_TEMPLATES_DIR / "print" / "_design.css").read_text(encoding="utf-8")
    assert "{%" not in icerik and "{{" not in icerik
    assert "--pr-ink" in icerik  # token seti yerinde


def test_sube_sirasi_turk_alfabesine_gore() -> None:
    """R2k/R4 şube sayfaları Türk alfabesi sırasında dizilir (kod noktası DEĞİL).

    Şube harfi artık ASCII'ye katlanmadığı için (gerçek e-Okul verisinde hem
    10/I hem 10/İ şubesi var), ham `str` karşılaştırması Ç/Ğ/İ/Ö/Ş/Ü'yü 'Z'den
    sonraya atıyordu: 10/Ç ve 10/İ evrakın sonuna düşer, 10/I ile 10/İ iki uca
    ayrılırdı.
    """
    etiketler = ["10/Z", "10/İ", "10/I", "10/Ç", "10/C", "9/B", "9/A", "11/A"]
    assert sorted(etiketler, key=reports.class_label_sort_key) == [
        "9/A",
        "9/B",
        "10/C",
        "10/Ç",
        "10/I",
        "10/İ",
        "10/Z",
        "11/A",
    ]


def test_sube_sirasi_seviye_sayisal_kalir() -> None:
    """Seviye sıralaması alfabetik DEĞİL sayısaldır (9 < 10 < 11) — davranış korundu."""
    assert sorted(["10/A", "9/A", "11/A"], key=reports.class_label_sort_key) == [
        "9/A",
        "10/A",
        "11/A",
    ]
    # Sayısal olmayan başlık (Hazırlık) daima sona düşer.
    assert sorted(["Hazırlık/A", "9/A"], key=reports.class_label_sort_key) == [
        "9/A",
        "Hazırlık/A",
    ]


def test_salon_sirasi_turk_alfabesine_gore() -> None:
    """Salon sayfaları Türk alfabesi sırasında dizilir (R1/R2/R3/R5/R6/R7/R9).

    Şube derslikleri `section_room_name` ile adlandırılır ve şube harfi artık
    ASCII'ye katlanmadığı için ('10/İ Dersliği') ham str karşılaştırması
    Ç/Ğ/İ/Ö/Ş/Ü'yü 'Z'den sonraya atıyordu: 10/I ile 10/İ dersliğinin sayfaları
    basılı evrakın iki ucuna düşerdi.
    """
    salonlar = [
        "10/Z Dersliği",
        "10/İ Dersliği",
        "10/I Dersliği",
        "10/Ş Dersliği",
        "10/S Dersliği",
        "10/Ç Dersliği",
        "10/C Dersliği",
    ]
    assert sorted(salonlar, key=reports.room_name_sort_key) == [
        "10/C Dersliği",
        "10/Ç Dersliği",
        "10/I Dersliği",
        "10/İ Dersliği",
        "10/S Dersliği",
        "10/Ş Dersliği",
        "10/Z Dersliği",
    ]


def test_salon_sirasi_serbest_metinde_cokmez() -> None:
    """Kelebek düzende salon adı serbest metindir — anahtar tip hatası vermemeli.

    Rakamla başlayan ve harfle başlayan adlar aynı listede karışabilir; anahtar
    tek biçimli demet üretmezse `sorted()` TypeError ile PDF üretimini düşürür.
    """
    karisik = ["A-101", "3 Nolu Salon", "Çok Amaçlı Salon", "10/İ Dersliği", "Fizik Lab"]
    sirali = sorted(karisik, key=reports.room_name_sort_key)  # çökmemeli
    assert set(sirali) == set(karisik)
    # Ayraç (boşluk/tire) harflerden ÖNCE gelir: 'A Salonu' < 'AB Salonu'.
    assert sorted(["AB Salonu", "A Salonu"], key=reports.room_name_sort_key) == [
        "A Salonu",
        "AB Salonu",
    ]
